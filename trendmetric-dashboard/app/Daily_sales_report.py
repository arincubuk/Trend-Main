import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import snowflake.connector

# ----------------------------------------------------------------
# 1) Configuration
# ----------------------------------------------------------------
OUTPUT_FOLDER = "/Users/arincubuk/Library/CloudStorage/OneDrive-furniq.co.uk/MACY'S SALES"
TEMPLATES_FOLDER = "templates"

# Correctly joined relative paths
master_file_path = os.path.join(TEMPLATES_FOLDER, "Updated_Furniq_Master_file.xlsx")
template_path = os.path.join(TEMPLATES_FOLDER, "Daily Sales Template.xlsx")

SALES_STATUSES = [
    "RECEIVED",
    "SHIPPED",
    "SHIPPING",
    "STAGING",
    "INCIDENT_OPEN",
    "WAITING_ACCEPTANCE",
    "CLOSED"
]
RETURN_STATUS = ["REFUNDED"]



# ----------------------------------------------------------------
# 2) Connect to Snowflake and Pull Data
# ----------------------------------------------------------------
conn = snowflake.connector.connect(
    user="ARINCUBUK",
    password="Shearling2024$",
    account="db19901.europe-west2.gcp",
    warehouse="MIRAKL",
    database="PRECOG",
    schema="MIRAKL_MACY_S_MIRAKL_MIRAKL",
    role="ACCOUNTADMIN"
)

cur = conn.cursor()

sql_query = """
    SELECT
        CAST(CREATED_DATE_ORDERS_ORDER_LINES AS DATE) AS "CREATED_DATE_ORDERS_ORDER_LINES",
        "OFFER_SKU_ORDERS_ORDER_LINES",
        "ORDER_LINE_STATE_ORDERS_ORDER_LINES",
        "QUANTITY_ORDERS_ORDER_LINES",
        "TOTAL_PRICE_ORDERS_ORDER_LINES",
        "ORDERS_PRECOG_KEY"
    FROM ORDERS_ORDER_LINES_NEW;
"""

cur.execute(sql_query)
rows = cur.fetchall()
col_names = [desc[0] for desc in cur.description]
cur.close()
conn.close()

# Create a DataFrame
df_snowflake = pd.DataFrame(rows, columns=col_names)
print("Columns in df_snowflake:", df_snowflake.columns.tolist())

# ----------------------------------------------------------------
# 3) Process Snowflake Data
# ----------------------------------------------------------------
# Convert 'CREATED_DATE_ORDERS_ORDER_LINES' to datetime and remove timezone
df_snowflake["CREATED_DATE_ORDERS_ORDER_LINES"] = pd.to_datetime(
    df_snowflake["CREATED_DATE_ORDERS_ORDER_LINES"], errors="coerce"
).dt.tz_localize(None)  # Remove timezone
df_snowflake.dropna(subset=["CREATED_DATE_ORDERS_ORDER_LINES"], inplace=True)

# ----------------------------------------------------------------
# 4) Load Master File for Product Data
# ----------------------------------------------------------------
master_df = pd.read_excel(master_file_path, engine="openpyxl")

# Merge Snowflake data with Master file (bringing in 'Product' and other columns)
df_merged = df_snowflake.merge(
    master_df,
    left_on="OFFER_SKU_ORDERS_ORDER_LINES",
    right_on="Variant SKU",  # Adjust if your master file uses a different column
    how="left"
)

# ----------------------------------------------------------------
# 5) Yesterday’s Sales
# ----------------------------------------------------------------
yesterday = datetime.now().date() - timedelta(days=1)

yesterdays_sales = df_merged[
    (df_merged["CREATED_DATE_ORDERS_ORDER_LINES"].dt.date == yesterday)
    & (df_merged["ORDER_LINE_STATE_ORDERS_ORDER_LINES"].isin(SALES_STATUSES))
]
# Sort by SKU (OFFER_SKU_ORDERS_ORDER_LINES)
yesterdays_sales = yesterdays_sales.sort_values(by="OFFER_SKU_ORDERS_ORDER_LINES", ascending=True)

yesterdays_sales_summary = yesterdays_sales[
    [
        "CREATED_DATE_ORDERS_ORDER_LINES",
        "ORDERS_PRECOG_KEY",
        "OFFER_SKU_ORDERS_ORDER_LINES",
        "QUANTITY_ORDERS_ORDER_LINES",
        "ORDER_LINE_STATE_ORDERS_ORDER_LINES",
        "TOTAL_PRICE_ORDERS_ORDER_LINES",
    ]
]

daily_category_totals = yesterdays_sales.groupby("Product").agg(
    {"QUANTITY_ORDERS_ORDER_LINES": "sum", "TOTAL_PRICE_ORDERS_ORDER_LINES": "sum"}
).reset_index()
daily_total_qty = yesterdays_sales["QUANTITY_ORDERS_ORDER_LINES"].sum()
daily_total_amt = yesterdays_sales["TOTAL_PRICE_ORDERS_ORDER_LINES"].sum()

# ----------------------------------------------------------------
# 6) Load Excel Template
# ----------------------------------------------------------------
template_wb = load_workbook(template_path)
template_ws = template_wb.active

# Clear Old Data in Template (Except Certain Cells)
start_row = 3
for row in template_ws.iter_rows(
    min_row=start_row,
    max_row=template_ws.max_row,
    min_col=1,
    max_col=template_ws.max_column
):
    for cell in row:
        if cell.coordinate not in [
            "G3", "G4", "G11", "G12", "G13", "G14", "G15", "G16",
            "G17", "G18", "G19", "G20", "G21", "G22", "G23", 'G24', 'G25', 'G26', 'G27', 'G28', 'G29', 'G30', 'G31', 'G32', 'G33', 'G34', 'G35',
            "H11", "I11", "J11", "K11", "L11", "J3", "J4", "J5",
            "J6", "J7", "J8", "J9", "M11", 'N2', 'N3', 'N4', 'N5', 'N6', 'N7', 'N8', 'O3', 'O4', 'O5', 'O6', 'O7', 'O8', 'N11', 'O11', 'P11', 'Q11', 'R11',
            'H5', 'H6', 'H7', 'P1'
        ]:
            cell.value = None

# ----------------------------------------------------------------
# 7) Populate Yesterday’s Sales in Template
# ----------------------------------------------------------------
for r_idx, row_data in enumerate(
    dataframe_to_rows(yesterdays_sales_summary, index=False, header=False),
    start=start_row
):
    for c_idx, value in enumerate(row_data, start=1):
        template_ws.cell(row=r_idx, column=c_idx, value=value)

template_ws["H3"] = daily_total_qty
template_ws["I3"] = daily_total_amt
# ----------------------------------------------------------------
# 7.1) Populate Category Counts for Yesterday's Sales
# ----------------------------------------------------------------
# Define categories and their corresponding Excel cell locations
categories = {
    "SHEARLING JACKET": "K3",
    "NAPPA JACKET": "K4",
    "TRENCH COAT": "K5",
    "BAG": "K6",
    "SLIPPER/BOOT": "K7",
}

# Define red fill for conditional formatting
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for category, cell in categories.items():
    # Filter sales data for the current category
    category_sales = yesterdays_sales[yesterdays_sales["Product"] == category]

    # Calculate the average price
    if not category_sales.empty:
        avg_price = category_sales["TOTAL_PRICE_ORDERS_ORDER_LINES"].sum() / category_sales["QUANTITY_ORDERS_ORDER_LINES"].sum()
    else:
        avg_price = 0  # If no sales, default to 0

    # Update the Excel cell with the calculated average price
    template_ws[cell] = avg_price

    # Apply conditional formatting for SHEARLING JACKET and NAPPA JACKET
    if (category == "SHEARLING JACKET" and avg_price < 400) or (category == "NAPPA JACKET" and avg_price < 180):
        template_ws[cell].fill = red_fill  # Apply red background

# ----------------------------------------------------------------
# ----------------------------------------------------------------
# 7.2) Calculate and Populate Time-Based Percentages for Yesterday’s Sales
# ----------------------------------------------------------------

# Define new time intervals
morning_start, morning_end = 4, 12    # 4:00 AM to 11:59 AM
afternoon_start, afternoon_end = 12, 18  # 12:00 PM to 5:59 PM
evening_start_1, evening_end_1 = 18, 24  # 6:00 PM to 11:59 PM
evening_start_2, evening_end_2 = 0, 4    # 12:00 AM to 3:59 AM (same day)

# Extract the hour safely using .loc
yesterdays_sales = yesterdays_sales.copy()  # Ensure we're working on a new DataFrame, not a slice
yesterdays_sales.loc[:, "hour"] = yesterdays_sales["CREATED_DATE_ORDERS_ORDER_LINES"].dt.hour

# Calculate total quantity for yesterday
total_yesterday_quantity = yesterdays_sales["QUANTITY_ORDERS_ORDER_LINES"].sum()

# Calculate percentages for each time period
morning_quantity = yesterdays_sales[
    (yesterdays_sales["hour"] >= morning_start) & (yesterdays_sales["hour"] < morning_end)
]["QUANTITY_ORDERS_ORDER_LINES"].sum()

afternoon_quantity = yesterdays_sales[
    (yesterdays_sales["hour"] >= afternoon_start) & (yesterdays_sales["hour"] < afternoon_end)
]["QUANTITY_ORDERS_ORDER_LINES"].sum()

# Evening now includes both:
# 1. 6:00 PM to 11:59 PM
# 2. 12:00 AM to 3:59 AM (same-day early morning sales)
evening_quantity = yesterdays_sales[
    ((yesterdays_sales["hour"] >= evening_start_1) & (yesterdays_sales["hour"] < evening_end_1)) |
    ((yesterdays_sales["hour"] >= evening_start_2) & (yesterdays_sales["hour"] < evening_end_2))
]["QUANTITY_ORDERS_ORDER_LINES"].sum()

# Ensure no division by zero
morning_percentage = (morning_quantity / total_yesterday_quantity * 100) if total_yesterday_quantity > 0 else 0
afternoon_percentage = (afternoon_quantity / total_yesterday_quantity * 100) if total_yesterday_quantity > 0 else 0
evening_percentage = (evening_quantity / total_yesterday_quantity * 100) if total_yesterday_quantity > 0 else 0

# Populate Excel template for yesterday’s sales (Converted to Decimal %)
template_ws["I5"] = morning_percentage / 100  # Morning %
template_ws["I6"] = afternoon_percentage / 100  # Afternoon %
template_ws["I7"] = evening_percentage / 100  # Evening %
# ----------------------------------------------------------------
# 7.3) Calculate Top 5 SKUs by Quantity for Yesterday’s Sales
# ----------------------------------------------------------------

# Group by Master SKU and calculate total quantity for each SKU
top_skus = (
    yesterdays_sales.groupby("Master SKU")["QUANTITY_ORDERS_ORDER_LINES"]
    .sum()
    .reset_index()
    .sort_values(by="QUANTITY_ORDERS_ORDER_LINES", ascending=False)
    .head(5)  # Select the top 5 SKUs
)

# Populate the top 5 SKUs into the Excel template
for idx, row in enumerate(top_skus.itertuples(index=False), start=3):  # Start at row 3
    template_ws[f"L{idx}"] = row[0]  # Master SKU
    template_ws[f"M{idx}"] = row[1]  # Quantity

# ----------------------------------------------------------------
# 8) Monthly Summaries
# ----------------------------------------------------------------
months = pd.date_range(start="2024-01-01 00:00:00", end=datetime.now(), freq="MS").strftime("%B %Y").tolist()
row_idx = 12

for month_label in months:
    month_start = pd.to_datetime(month_label)
    month_end = month_start + pd.offsets.MonthEnd(0)

    month_data = df_merged[
        (df_merged["CREATED_DATE_ORDERS_ORDER_LINES"] >= month_start)
        & (df_merged["CREATED_DATE_ORDERS_ORDER_LINES"] <= month_end)
    ]

    # Sales and Returns
    sales_subset = month_data[month_data["ORDER_LINE_STATE_ORDERS_ORDER_LINES"].isin(SALES_STATUSES)]
    returns_subset = month_data[month_data["ORDER_LINE_STATE_ORDERS_ORDER_LINES"].isin(RETURN_STATUS)]

    net_value = sales_subset["TOTAL_PRICE_ORDERS_ORDER_LINES"].sum()
    sales_quantity = sales_subset["QUANTITY_ORDERS_ORDER_LINES"].sum()
    returns_quantity = returns_subset["QUANTITY_ORDERS_ORDER_LINES"].sum()
    gross_quantity = sales_quantity + returns_quantity

    return_percentage = (returns_quantity / gross_quantity * 100) if gross_quantity > 0 else 0
    return_percentage = round(return_percentage, 2)

    gross_value = month_data["TOTAL_PRICE_ORDERS_ORDER_LINES"].sum()

    template_ws[f"I{row_idx}"] = net_value
    template_ws[f"O{row_idx}"] = gross_value
    template_ws[f"M{row_idx}"] = return_percentage
    template_ws[f"N{row_idx}"] = gross_quantity
    row_idx += 1
# ----------------------------------------------------------------
# 8.1) Populate Category Counts for Monthly Summaries
# ----------------------------------------------------------------
row_idx_shearling = 12  # Start row for SHEARLING JACKET (J column)
row_idx_nappa = 12  # Start row for NAPPA JACKET (K column)
row_idx_trench = 12  # Start row for TRENCH COAT JACKET (L column)

for month_label in months:
    month_start = pd.to_datetime(month_label)
    month_end = month_start + pd.offsets.MonthEnd(0)

    # Filter data for the current month
    month_data = df_merged[
        (df_merged["CREATED_DATE_ORDERS_ORDER_LINES"] >= month_start)
        & (df_merged["CREATED_DATE_ORDERS_ORDER_LINES"] <= month_end)
    ]
    # Calculate total sales quantity for the month
    total_month_quantity = month_data["QUANTITY_ORDERS_ORDER_LINES"].sum()

    # Calculate SHEARLING JACKET count for the month
    shearling_count = month_data[month_data["Product"] == "SHEARLING JACKET"]["QUANTITY_ORDERS_ORDER_LINES"].sum()
    shearling_percentage = (shearling_count / total_month_quantity * 100) if total_month_quantity > 0 else 0
    shearling_percentage = round(shearling_percentage, 2)  # Round to 2 decimal places
    template_ws[f"J{row_idx_shearling}"] = shearling_percentage / 100

    # Calculate NAPPA JACKET count for the month
    nappa_count = month_data[month_data["Product"] == "NAPPA JACKET"]["QUANTITY_ORDERS_ORDER_LINES"].sum()
    nappa_percentage = (nappa_count / total_month_quantity * 100) if total_month_quantity > 0 else 0
    nappa_percentage = round(nappa_percentage, 2)  # Round to 2 decimal places
    template_ws[f"K{row_idx_nappa}"] = nappa_percentage / 100  # Update column K

    # Calculate Trench JACKET count for the month
    trench_count = month_data[month_data["Product"] == "TRENCH COAT"]["QUANTITY_ORDERS_ORDER_LINES"].sum()
    trench_percentage = (trench_count / total_month_quantity * 100) if total_month_quantity > 0 else 0
    trench_percentage = round(trench_percentage, 2)  # Round to 2 decimal places
    template_ws[f"L{row_idx_nappa}"] = trench_percentage / 100  # Update column L

    # Increment row indices for the next month
    row_idx_shearling += 1
    row_idx_nappa += 1
    row_idx_trench += 1
# ----------------------------------------------------------------
# 8.2) Populate Monthly Sales Quantity in Column H (Starting H12)
# ----------------------------------------------------------------
row_idx_quantity = 12  # Start row for sales quantity in Column H

for month_label in months:
    month_start = pd.to_datetime(month_label)
    month_end = month_start + pd.offsets.MonthEnd(0)
    # Filter data for the current month and only include Sales statuses
    month_sales_data = df_merged[
        (df_merged["CREATED_DATE_ORDERS_ORDER_LINES"] >= month_start)
        & (df_merged["CREATED_DATE_ORDERS_ORDER_LINES"] <= month_end)
        & (df_merged["ORDER_LINE_STATE_ORDERS_ORDER_LINES"].isin(SALES_STATUSES))
        ]

    # Calculate total sales quantity for the month
    sales_quantity = month_sales_data["QUANTITY_ORDERS_ORDER_LINES"].sum()

    # Populate the sales quantity in Column H for the current month
    template_ws[f"H{row_idx_quantity}"] = sales_quantity

    # Increment the row index for the next month
    row_idx_quantity += 1
# ----------------------------------------------------------------
# 8.4) Calculate and Populate Time-Based Percentages for Monthly Sales
# ----------------------------------------------------------------
row_idx_morning = 12  # Start row for Morning % (Column P)
row_idx_afternoon = 12  # Start row for Afternoon % (Column Q)
row_idx_evening = 12  # Start row for Evening % (Column R)

for month_label in months:
    month_start = pd.to_datetime(month_label)
    month_end = month_start + pd.offsets.MonthEnd(0)

    # Filter data for the current month
    month_data = df_merged[
        (df_merged["CREATED_DATE_ORDERS_ORDER_LINES"] >= month_start)
        & (df_merged["CREATED_DATE_ORDERS_ORDER_LINES"] <= month_end)
    ].copy()  # Copy the filtered DataFrame

    # Extract the hour safely
    month_data.loc[:, "hour"] = month_data["CREATED_DATE_ORDERS_ORDER_LINES"].dt.hour

    # Calculate total quantity for the month
    total_month_quantity = month_data["QUANTITY_ORDERS_ORDER_LINES"].sum()

    # Calculate percentages for each time period
    morning_quantity = month_data[
        (month_data["hour"] >= morning_start) & (month_data["hour"] < morning_end)
    ]["QUANTITY_ORDERS_ORDER_LINES"].sum()
    morning_percentage = (morning_quantity / total_month_quantity * 100) if total_month_quantity > 0 else 0

    afternoon_quantity = month_data[
        (month_data["hour"] >= afternoon_start) & (month_data["hour"] < afternoon_end)
    ]["QUANTITY_ORDERS_ORDER_LINES"].sum()
    afternoon_percentage = (afternoon_quantity / total_month_quantity * 100) if total_month_quantity > 0 else 0

    evening_quantity = month_data[
        ((month_data["hour"] >= evening_start_1) & (month_data["hour"] < evening_end_1)) |
        ((month_data["hour"] >= evening_start_2) & (month_data["hour"] < evening_end_2))
    ]["QUANTITY_ORDERS_ORDER_LINES"].sum()
    evening_percentage = (evening_quantity / total_month_quantity * 100) if total_month_quantity > 0 else 0

    # Populate Excel template for the current month
    # Populate Excel template for monthly sales percentages
    template_ws[f"P{row_idx_morning}"] = morning_percentage / 100  # Morning %
    template_ws[f"Q{row_idx_afternoon}"] = afternoon_percentage / 100  # Afternoon %
    template_ws[f"R{row_idx_evening}"] = evening_percentage / 100  # Evening %

    # Increment row indices for the next month
    row_idx_morning += 1
    row_idx_afternoon += 1
    row_idx_evening += 1
# ----------------------------------------------------------------
# 9) Save Populated Template
# ----------------------------------------------------------------
current_date_str = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
output_excel = os.path.join(OUTPUT_FOLDER, f"Sales Report {current_date_str}.xlsx")

print("Attempting to save Excel file to:", output_excel)
try:
    template_wb.save(output_excel)
    print("Excel file saved successfully.")
except Exception as e:
    print("Error saving the Excel file:", e)

from io import BytesIO
def generate_daily_sales_report():
    # all your existing logic remains unchanged...

    output = BytesIO()
    template_wb.save(output)
    output.seek(0)
    return output, f"Sales Report {current_date_str}.xlsx"
